import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 建立資料
data = [
    ["固定金額", ""],
    ["薪水", "500 元 / 小時"],
    ["額外費用", ""],
    ["ChatGPT Plus", "20 美金 / 月"],
    ["撰寫 API 所需 Token 費用", "依使用數量計算"],
    ["APP 後端 API 上雲費用", "依使用數量計算"],
    ["研究所產生的額外支出", "依使用數量計算"]
]

# 建立 DataFrame
df = pd.DataFrame(data, columns=["項目", "內容"])

# 建立 Excel 工作簿
wb = Workbook()
ws = wb.active
ws.title = "費用表"

# 寫入表頭
for col_num, column_title in enumerate(df.columns, 1):
    cell = ws.cell(row=1, column=col_num, value=column_title)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    cell.border = thin_border

# 寫入資料
for row_num, row_data in enumerate(data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        if row_data[1] == "":  # 標題行加底色
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# 調整欄寬
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 25

# 儲存檔案
file_path = "/mnt/data/費用表.xlsx"
wb.save(file_path)

file_path
